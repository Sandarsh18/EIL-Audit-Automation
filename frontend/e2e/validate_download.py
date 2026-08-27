import sys
import openpyxl
import hashlib

def hash_file(filepath: str) -> str:
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        while chunk := f.read(8192):
            h.update(chunk)
    return h.hexdigest()

if __name__ == "__main__":
    final_path = sys.argv[1]
    e3_path = sys.argv[2]
    
    out_hash = hash_file(final_path)
    e3_hash = hash_file(e3_path)
    
    assert out_hash != e3_hash, "Downloaded file matches source Excel 3!"
    
    wb = openpyxl.load_workbook(final_path, data_only=True)
    assert 'ConsolidatedMHrequirementApr25' in wb.sheetnames
    sheet = wb['ConsolidatedMHrequirementApr25']
    
    # Read all data
    data = list(sheet.iter_rows(values_only=True))
    
    # Ensure B269 is in the rows
    found = False
    for row in data:
        # Check if the row contains B269
        if row and 'B269' in str(row):
            found = True
            
        # Check for blank columns (e.g., column index 2 and 8+ which are unmapped typically)
        # We just need to assert there are some blank columns!
        if row and row[0] is not None and str(row[0]) != 'Job No':
            assert row[2] is None or row[2] == '', f"Expected blank column for unmapped data but found {row[2]}"
            
    assert found, "Approved job B269 not found in output!"
    assert len(data) < 10, "Data contains too many rows, original data not deleted!"
    
    print("VALIDATION SUCCESS")
