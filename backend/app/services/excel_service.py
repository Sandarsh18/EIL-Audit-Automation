import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from app.schemas.workbook import WorkbookMetadata, SheetMetadata, ColumnMetadata, SheetSummary
from fastapi import HTTPException
import numpy as np

class ExcelService:
    @staticmethod
    def detect_header_row(file_path: str, sheet_name: str) -> int:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
            df = df.fillna("")
            
            best_row_idx = 0
            max_score = -1
            
            canonical_terms = ["job no", "job number", "orders", "exp.", "inspn", "others", "total", "ocs"]
            
            for idx, row in df.iterrows():
                row_str = [str(x).strip().lower() for x in row]
                score = sum(1 for term in canonical_terms if any(term in x for x in row_str))
                
                str_count = sum(1 for x in row_str if x)
                total_score = score * 10 + str_count
                
                if total_score > max_score:
                    max_score = total_score
                    best_row_idx = idx
                    
            return best_row_idx
        except Exception:
            return 0

    @staticmethod
    def get_workbook_metadata(file_path: str, file_id: str, filename: str, workbook_type: str) -> WorkbookMetadata:
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found on disk")
        
        try:
            size = os.path.getsize(file_path)
            
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            
            # Extract basic sheet metadata for candidates
            sheet_summaries = []
            try:
                # Read first 20 rows of all sheets to detect headers efficiently
                dfs = pd.read_excel(file_path, sheet_name=None, nrows=20)
                
                for sheet_name in sheet_names:
                    ws = wb[sheet_name]
                    max_row = ws.max_row if ws.max_row is not None else 0
                    
                    df = dfs.get(sheet_name)
                    if df is None or df.empty:
                        sheet_summaries.append(SheetSummary(
                            name=sheet_name, is_candidate=False, schema_type="unknown", row_count=max_row, columns=[]
                        ))
                        continue
                        
                    best_row_idx = 0
                    max_non_null = 0
                    for idx, row in df.iterrows():
                        non_null_count = row.notna().sum()
                        if non_null_count > max_non_null:
                            max_non_null = non_null_count
                            best_row_idx = idx
                            
                    header_row = [str(x) for x in df.iloc[best_row_idx].tolist()]
                    header_row_lower = [h.lower() for h in header_row]
                    
                    has_job_no = any("job" in h and "no" in h for h in header_row_lower)
                    has_orders_for = any("orders for" in h for h in header_row_lower)
                    has_ocs_done = any("ocs done" in h for h in header_row_lower)
                    
                    is_candidate = has_job_no
                    schema = "unknown"
                    if has_orders_for or has_ocs_done:
                        schema = "modern"
                    elif has_job_no:
                        schema = "legacy"
                        
                    sheet_summaries.append(SheetSummary(
                        name=sheet_name,
                        is_candidate=is_candidate,
                        schema_type=schema,
                        row_count=max_row,
                        columns=header_row
                    ))
            except Exception as inner_e:
                # Fallback if pandas read_excel(sheet_name=None) fails
                print(f"Warning: Failed to parse sheet summaries: {inner_e}")
                pass
                
            wb.close()
            
            return WorkbookMetadata(
                file_id=file_id,
                workbook_type=workbook_type,
                filename=filename,
                size=size,
                sheets=sheet_names,
                sheet_summaries=sheet_summaries if sheet_summaries else None
            )
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to open workbook: {str(e)}")

    @staticmethod
    def get_sheet_metadata(file_path: str, sheet_name: str, preview_rows: int = 50, workbook_type: str = "excel3") -> SheetMetadata:
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found on disk")

        try:
            # Read only the first N rows for preview + some extra for type inference
            # We'll read the whole sheet for row/column counts, but in a real enterprise app 
            # we might optimize this. For Phase 1, we read it defensively.
            # Detect the correct header row first
            header_idx = ExcelService.detect_header_row(file_path, sheet_name)
            
            # Using str to avoid pandas dropping/converting unexpected types too aggressively,
            # but we need to infer types.
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx, dtype=object)
            
            # Ensure column names are unique strings
            df.columns = [str(c) if pd.notna(c) and str(c).strip() != "" else f"Unnamed: {i}" for i, c in enumerate(df.columns)]

            row_count = len(df)
            column_count = len(df.columns)
            
            columns_meta: List[ColumnMetadata] = []
            
            for i, col_name in enumerate(df.columns):
                # Infer type
                col_data = df[col_name].dropna()
                if len(col_data) == 0:
                    inferred_type = "blank"
                else:
                    # Let's try to infer type by trying to convert
                    types_found = set()
                    for val in col_data:
                        if isinstance(val, (int, float)):
                            types_found.add("number")
                        elif isinstance(val, pd.Timestamp) or isinstance(val, datetime):
                            types_found.add("date")
                        elif isinstance(val, bool):
                            types_found.add("boolean")
                        elif isinstance(val, str):
                            types_found.add("text")
                        else:
                            types_found.add("mixed")
                    
                    if len(types_found) == 1:
                        inferred_type = types_found.pop()
                    elif len(types_found) > 1:
                        if types_found.issubset({"number", "text"}):
                            inferred_type = "mixed" # text and numbers
                        else:
                            inferred_type = "mixed"
                    else:
                        inferred_type = "mixed"
                
                import re
                canonical_name = None
                col_name_lower = str(col_name).lower()
                
                if workbook_type == "excel1":
                    if re.search(r"job.*no", col_name_lower): canonical_name = "JOB_NUMBER"
                    elif re.search(r"balance", col_name_lower): canonical_name = "BALANCE_QUANTITY"
                    elif re.search(r"ocs.*date", col_name_lower): canonical_name = "OCS_DATE"
                elif workbook_type == "excel2":
                    if re.search(r"job.*no", col_name_lower): canonical_name = "JOB_NUMBER"
                    elif re.search(r"attended.*from", col_name_lower): canonical_name = "INSPECTION_FROM"
                    elif re.search(r"attended.*upto", col_name_lower): canonical_name = "INSPECTION_UPTO"
                    elif re.search(r"date.*received", col_name_lower): canonical_name = "DATE_RECEIVED"
                    elif re.search(r"qap.*appl", col_name_lower): canonical_name = "QAP_APPL"
                    elif re.search(r"no.*days|working.*days", col_name_lower): canonical_name = "NO_OF_WORKING_DAYS"
                elif workbook_type == "excel3":
                    if re.search(r"job.*no", col_name_lower): canonical_name = "JOB_NUMBER"
                    elif re.search(r"running.*order|no.*orders", col_name_lower): canonical_name = "RUNNING_ORDERS"
                    elif re.search(r"orders.*for", col_name_lower): canonical_name = "ORDERS_FOR_FD_FOLLOWUP"
                    elif re.search(r"ocs.*done", col_name_lower): canonical_name = "OCS_DONE"
                    elif re.search(r"^exp", col_name_lower): canonical_name = "EXPEDITING"
                    elif re.search(r"^insp", col_name_lower): canonical_name = "INSPECTION_SOURCE"
                    elif re.search(r"others", col_name_lower): canonical_name = "OTHERS"
                    elif re.search(r"total", col_name_lower): canonical_name = "TOTAL"
                
                columns_meta.append(ColumnMetadata(
                    name=str(col_name),
                    index=i + 1,
                    data_type=inferred_type,
                    canonical=canonical_name
                ))

            # Preview data: replace NaNs with None for JSON serialization
            preview_df = df.head(preview_rows).replace({np.nan: None})
            preview_data = preview_df.to_dict(orient="records")

            return SheetMetadata(
                sheet_name=sheet_name,
                row_count=row_count,
                column_count=column_count,
                columns=columns_meta,
                preview=preview_data
            )
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to inspect sheet: {str(e)}")
