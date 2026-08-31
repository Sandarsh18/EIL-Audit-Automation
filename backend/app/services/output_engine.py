import os
import shutil
import hashlib
import openpyxl
import copy
from datetime import datetime, timezone
from typing import List, Dict, Any, Tuple, Optional
from fastapi import HTTPException
from app.schemas.output import (
    ChangePlanCell, ChangePlan, OutputMetadata, OutputPlanRequest, OutputGenerateRequest, ActionType
)
from app.services.session_service import SessionService
from app.services.review_service import ReviewService
from app.services.matching_service import normalize_job_number
from openpyxl.utils import get_column_letter

class OutputEngine:
    
    @staticmethod
    def _hash_file(filepath: str) -> str:
        h = hashlib.sha256()
        with open(filepath, 'rb') as f:
            while chunk := f.read(8192):
                h.update(chunk)
        return h.hexdigest()
        
    @staticmethod
    def _find_header_row_and_col(sheet, header_name: str) -> Tuple[Optional[int], Optional[int]]:
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                if cell_value and str(cell_value).strip() == header_name:
                    return row_idx, col_idx
        return None, None

    @staticmethod
    def generate_change_plan(session_id: str, request: OutputPlanRequest) -> ChangePlan:
        session = SessionService.get_session(session_id)
        from app.config import UPLOAD_DIR
        e3_path = os.path.join(UPLOAD_DIR, f"{session.excel3_file_id}.xlsx")
        
        if not os.path.exists(e3_path):
            raise HTTPException(400, "Original Excel 3 not found on disk. Please upload it again.")
            
        wb = openpyxl.load_workbook(e3_path, data_only=False)
        reviews = ReviewService.get_reviews(session_id, request.job_numbers, request.evaluation_month)
        job_reviews = {r.job_number: r for r in reviews}
        
        mapping = session.mapping.excel3.columns.model_dump(exclude_none=True)
        target_sheet = session.mapping.excel3.sheet
        
        if target_sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(400, f"Mapped sheet {target_sheet} not found in Excel 3")
            
        sheet = wb[target_sheet]
        header_row, job_col_idx = OutputEngine._find_header_row_and_col(sheet, mapping.get('job_number'))
        
        if not header_row or not job_col_idx:
            wb.close()
            raise HTTPException(400, "Could not find mapped Job Number header in Excel 3 template")

        headers = {}
        for col_idx, cell in enumerate(sheet[header_row], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        cells_to_modify = []
        approved_jobs_included = 0
        blocked_jobs = []
        
        # We just generate the "planned" modifications to show what will be inserted
        for row_offset, job_no in enumerate(request.job_numbers, start=1):
            if job_no not in job_reviews:
                blocked_jobs.append(job_no)
                continue
                
            job_review = job_reviews[job_no]
            if job_review.status != "APPROVED":
                blocked_jobs.append(job_no)
                continue
                
            approved_jobs_included += 1
            virtual_row_index = header_row + row_offset
            
            for logical_field, header_name in mapping.items():
                if logical_field == 'job_number':
                    new_val = job_no
                elif logical_field == 'orders_for_fd':
                    new_val = getattr(job_review, 'fd', None)
                elif logical_field == 'total':
                    new_val = getattr(job_review, 'calculated_total', None)
                elif not hasattr(job_review, logical_field):
                    continue
                else:
                    new_val = getattr(job_review, logical_field, None)
                    
                col_idx = headers.get(header_name)
                if not col_idx:
                    continue
                    
                cells_to_modify.append(ChangePlanCell(
                    session_id=session_id,
                    job_number=job_no,
                    sheet_name=target_sheet,
                    row_index=virtual_row_index,
                    column_index=col_idx,
                    cell_address=f"{get_column_letter(col_idx)}{virtual_row_index}",
                    logical_field=logical_field,
                    old_value=None,
                    new_value=new_val,
                    action=ActionType.MODIFY_VALUE,
                    reason="Output generation",
                    source="Business Rule",
                    approval_status=job_review.status,
                    is_formula=False,
                    formula_overwrite_approved=False
                ))
                
            has_meeting = getattr(job_review, 'meeting', None) is not None
            if has_meeting:
                # Add virtual column for preview
                meeting_val = job_review.meeting
                cells_to_modify.append(ChangePlanCell(
                    session_id=session_id,
                    job_number=job_no,
                    sheet_name=target_sheet,
                    row_index=virtual_row_index,
                    column_index=999, # Virtual index
                    cell_address=f"CUSTOM_{virtual_row_index}",
                    logical_field="meeting",
                    old_value=None,
                    new_value=meeting_val,
                    action=ActionType.MODIFY_VALUE,
                    reason="Output generation",
                    source="Business Rule",
                    approval_status=job_review.status,
                    is_formula=False,
                    formula_overwrite_approved=False
                ))
                
        wb.close()
        
        return ChangePlan(
            cells_to_modify=cells_to_modify,
            cells_unchanged=0,
            blocked_jobs=blocked_jobs,
            approved_jobs_included=approved_jobs_included,
            formula_overwrites=0
        )

    @staticmethod
    def generate_output(session_id: str, request: OutputGenerateRequest) -> OutputMetadata:
        session = SessionService.get_session(session_id)
        from app.config import UPLOAD_DIR
        e3_path = os.path.join(UPLOAD_DIR, f"{session.excel3_file_id}.xlsx")
        
        if not os.path.exists(e3_path):
            raise HTTPException(400, "Original Excel 3 not found.")
            
        orig_hash = OutputEngine._hash_file(e3_path)
        
        reviews = ReviewService.get_reviews(session_id, request.job_numbers, request.evaluation_month)
        job_reviews = {r.job_number: r for r in reviews}
        
        mapping = session.mapping.excel3.columns.model_dump(exclude_none=True)
        target_sheet = session.mapping.excel3.sheet
        
        # Inject Meeting as a custom column if there is meeting data
        has_meeting = any(getattr(jr, 'meeting', None) is not None for jr in job_reviews.values())
        if has_meeting:
            from app.schemas.output import CustomColumnData
            meeting_cc = CustomColumnData(
                heading="Meeting",
                data={j: getattr(jr, 'meeting') for j, jr in job_reviews.items() if getattr(jr, 'meeting', None) is not None}
            )
            if not request.custom_columns:
                request.custom_columns = []
            if not any(cc.heading.lower() == "meeting" for cc in request.custom_columns):
                request.custom_columns.insert(0, meeting_cc)
        
        working_dir = os.path.join(os.path.dirname(UPLOAD_DIR), "working", session_id)
        os.makedirs(working_dir, exist_ok=True)
        tmp_out = os.path.join(working_dir, "CONSOLIDATED_Manhour_Automated.xlsx")
        
        shutil.copy2(e3_path, tmp_out)
        
        wb = openpyxl.load_workbook(tmp_out, keep_vba=True)
        if target_sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(400, f"Mapped sheet {target_sheet} not found in Excel 3")
            
        # Delete all other sheets to ensure a clean customized workbook
        for sname in wb.sheetnames:
            if sname != target_sheet:
                del wb[sname]
                
        sheet = wb[target_sheet]
        header_row, job_col_idx = OutputEngine._find_header_row_and_col(sheet, mapping.get('job_number'))
        
        if not header_row or not job_col_idx:
            wb.close()
            raise HTTPException(400, "Could not find mapped Job Number header in Excel 3 template")
            
        import re
        from datetime import datetime
        
        eval_month_str = request.evaluation_month or getattr(session, 'evaluation_month', None)
        if eval_month_str:
            try:
                dt = datetime.strptime(eval_month_str, "%Y-%m")
                
                # Rename the sheet dynamically based on evaluation month (e.g., ConsolidatedMHrequirementAug26)
                sheet.title = f"ConsolidatedMHrequirement{dt.strftime('%b%y')}"
                
                formatted_month = dt.strftime("%b'%y")
                pattern = re.compile(r"for\s+[A-Za-z]{3}\s*'\d{2}")
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if pattern.search(cell.value):
                                cell.value = pattern.sub(f"for {formatted_month}", cell.value)
            except ValueError:
                pass

        # Map header text to column index
        headers = {}
        for col_idx, cell in enumerate(sheet[header_row], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
                
        # Insert Custom Columns headers if provided
        custom_column_indices = {}
        if request.custom_columns:
            tot_col_idx = headers.get(mapping.get('total'))
            num_cols = len(request.custom_columns)
            
            if tot_col_idx:
                sheet.insert_cols(tot_col_idx, amount=num_cols)
                start_col = tot_col_idx
                
                # Update headers map since columns shifted right
                for key, val in list(headers.items()):
                    if val >= tot_col_idx:
                        headers[key] = val + num_cols
            else:
                start_col = sheet.max_column + 1
                
            for i, cc in enumerate(request.custom_columns):
                col_idx = start_col + i
                sheet.cell(row=header_row, column=col_idx).value = cc.heading
                custom_column_indices[cc.heading] = col_idx
                # Apply style from previous header if possible
                prev_cell = sheet.cell(row=header_row, column=start_col - 1)
                if prev_cell.has_style:
                    sheet.cell(row=header_row, column=col_idx).font = copy.copy(prev_cell.font)
                    sheet.cell(row=header_row, column=col_idx).fill = copy.copy(prev_cell.fill)
                    sheet.cell(row=header_row, column=col_idx).border = copy.copy(prev_cell.border)
                    sheet.cell(row=header_row, column=col_idx).alignment = copy.copy(prev_cell.alignment)
                
        # 1. Capture styles of existing template rows (including Leave and Total if present)
        max_row = sheet.max_row
        leave_row_orig = None
        total_row_orig = None
        for r_idx in range(header_row + 1, max_row + 1):
            val = sheet.cell(row=r_idx, column=job_col_idx).value
            if val and isinstance(val, str):
                v_lower = str(val).lower().strip()
                if v_lower == "leave":
                    leave_row_orig = r_idx
                elif v_lower == "total":
                    total_row_orig = r_idx

        default_style_row = header_row + 1
        styles = {}
        if default_style_row <= max_row:
            for col_idx in range(1, sheet.max_column + 1):
                c = sheet.cell(row=default_style_row, column=col_idx)
                styles[col_idx] = {'font': copy.copy(c.font), 'border': copy.copy(c.border), 'fill': copy.copy(c.fill), 'number_format': c.number_format, 'alignment': copy.copy(c.alignment)}

        leave_styles = {}
        if leave_row_orig:
            for col_idx in range(1, sheet.max_column + 1):
                c = sheet.cell(row=leave_row_orig, column=col_idx)
                leave_styles[col_idx] = {'font': copy.copy(c.font), 'border': copy.copy(c.border), 'fill': copy.copy(c.fill), 'number_format': c.number_format, 'alignment': copy.copy(c.alignment)}

        total_styles = {}
        if total_row_orig:
            for col_idx in range(1, sheet.max_column + 1):
                c = sheet.cell(row=total_row_orig, column=col_idx)
                total_styles[col_idx] = {'font': copy.copy(c.font), 'border': copy.copy(c.border), 'fill': copy.copy(c.fill), 'number_format': c.number_format, 'alignment': copy.copy(c.alignment)}

        # Delete all existing data rows below the header row
        if max_row > header_row:
            sheet.delete_rows(header_row + 1, max_row - header_row)

        def apply_style(cell, style_dict):
            if not style_dict: return
            if style_dict.get('font'): cell.font = copy.copy(style_dict['font'])
            if style_dict.get('border'): cell.border = copy.copy(style_dict['border'])
            if style_dict.get('fill'): cell.fill = copy.copy(style_dict['fill'])
            if style_dict.get('number_format'): cell.number_format = style_dict['number_format']
            if style_dict.get('alignment'): cell.alignment = copy.copy(style_dict['alignment'])

        cells_modified = 0
        jobs_processed = 0
        jobs_blocked = 0
        
        # 2. Append new rows for approved jobs only
        first_job_row = header_row + 1
        current_row_idx = first_job_row
        for job_no in request.job_numbers:
            if job_no not in job_reviews or job_reviews[job_no].status != "APPROVED":
                jobs_blocked += 1
                continue
                
            job_review = job_reviews[job_no]
            
            # Set unmapped columns to blank/None implicitly by openpyxl cell creation or explicitly
            # For each column in the sheet, ensure it exists and is None
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=current_row_idx, column=col_idx).value = None
            
            for logical_field, header_name in mapping.items():
                if logical_field == 'job_number':
                    new_val = job_no
                elif logical_field == 'orders_for_fd':
                    new_val = getattr(job_review, 'fd', None)
                elif logical_field == 'total':
                    new_val = getattr(job_review, 'calculated_total', None)
                elif not hasattr(job_review, logical_field):
                    continue
                else:
                    new_val = getattr(job_review, logical_field, None)
                    
                col_idx = headers.get(header_name)
                if col_idx:
                    sheet.cell(row=current_row_idx, column=col_idx).value = new_val
                    cells_modified += 1
                    
            if request.custom_columns:
                for cc in request.custom_columns:
                    col_idx = custom_column_indices.get(cc.heading)
                    if col_idx:
                        sheet.cell(row=current_row_idx, column=col_idx).value = cc.data.get(job_no)
                        cells_modified += 1
                    
            # Apply default style to job row
            for col_idx in range(1, sheet.max_column + 1):
                apply_style(sheet.cell(row=current_row_idx, column=col_idx), styles.get(col_idx))
                    
            jobs_processed += 1
            current_row_idx += 1
            
        last_job_row = current_row_idx - 1
        
        # 3. Add Leave Row
        leave_row_idx = current_row_idx
        sheet.cell(row=leave_row_idx, column=job_col_idx).value = "Leave"
        for col_idx in range(1, sheet.max_column + 1):
            apply_style(sheet.cell(row=leave_row_idx, column=col_idx), leave_styles.get(col_idx) or styles.get(col_idx))
            
        # 4. Add Total Row
        current_row_idx += 1
        total_row_idx = current_row_idx
        sheet.cell(row=total_row_idx, column=job_col_idx).value = "Total"
        for col_idx in range(1, sheet.max_column + 1):
            apply_style(sheet.cell(row=total_row_idx, column=col_idx), total_styles.get(col_idx) or styles.get(col_idx))
            # Ensure the cell is blank unless it's the job_col_idx
            if col_idx != job_col_idx:
                sheet.cell(row=total_row_idx, column=col_idx).value = None
            
        # Identify columns to add SUM formula
        ro_col_idx = headers.get(mapping.get('running_orders'))
        tot_col_idx = headers.get(mapping.get('total'))
        
        sum_cols = []
        if ro_col_idx: sum_cols.append(ro_col_idx)
        if tot_col_idx: sum_cols.append(tot_col_idx)
        
        # Add SUM formulas
        for col_idx in sum_cols:
            col_letter = get_column_letter(col_idx)
            if last_job_row >= first_job_row:
                formula = f"=SUM({col_letter}{first_job_row}:{col_letter}{last_job_row})"
            else:
                formula = "=0"
            sheet.cell(row=total_row_idx, column=col_idx).value = formula
            cells_modified += 1
            
        wb.save(tmp_out)
        wb.close()
        
        orig_hash_after = OutputEngine._hash_file(e3_path)
        if orig_hash != orig_hash_after:
            os.remove(tmp_out)
            raise HTTPException(500, "FATAL: Original file was modified during output generation.")
            
        import uuid
        output_id = str(uuid.uuid4())
        SessionService.set_generated_output(session_id, tmp_out, output_id)
            
        return OutputMetadata(
            status="SUCCESS",
            filename="CONSOLIDATED_Manhour_Automated.xlsx",
            output_path=tmp_out,
            output_id=output_id,
            original_sha256=orig_hash,
            original_unchanged=True,
            jobs_processed=jobs_processed,
            jobs_blocked=jobs_blocked,
            cells_modified=cells_modified,
            unexpected_changes=0,
            formula_changes=0,
            unexpected_formula_changes=0,
            sheets_changed=[target_sheet]
        )
