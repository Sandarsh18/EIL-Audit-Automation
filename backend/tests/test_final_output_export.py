import pytest
from fastapi.testclient import TestClient
from app.main import app
import pandas as pd
import openpyxl
import os
from app.config import UPLOAD_DIR

client = TestClient(app)

def test_final_output_export():
    resp = client.post("/api/sessions")
    s_id = resp.json()["session_id"]
    
    # 1. Setup Excel 1
    df1 = pd.DataFrame({
        "Job No.": ["B269", "B224", "B285", "B302", "B378"],
        "Balance Quantity": [0, 10, 0, 0, 0],
        "OCS Date": ["01/08/2026", None, "05/08/2026", "10/08/2026", "15/08/2026"]
    })
    
    # 2. Setup Excel 2
    # B269: Inspection=2, Others=0.5
    # B224: Inspection=3, Others=12.5
    # B285: Inspection=1, Others=2
    # B302: Inspection=1, Others=3.5
    # B378: Inspection=1, Others=7.25
    df2 = pd.DataFrame({
        "JOB NO.": ["B269", "B224", "B285", "B302", "B378"],
        "FROM": ["2026-08-01", "2026-08-10", "2026-08-01", "2026-08-01", "2026-08-01"],
        "UPTO": ["2026-08-02", "2026-08-12", "2026-08-01", "2026-08-01", "2026-08-01"],
        "NO. OF DAYS": [0.5, None, 2, 3.5, 7.25],
        "DATE RECEIVED": ["2026-06-15", "2026-05-10", "2026-06-15", "2026-06-15", "2026-06-15"],
        "QAP APPL.": [None, 12.5, None, None, None]
    })
    
    # 3. Setup Excel 3 with Title "Consolidated man hour requirement for Mar'26"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="Consolidated man hour requirement for Mar'26")
    ws.cell(row=2, column=5, value="MH to be released for Mar'26")
    
    headers = ["Job No.", "Running Orders", "Orders for FD", "OCS done", "Exp.", "Inspn", "Others", "Total"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=h)
        
    os.makedirs("tests/fixtures", exist_ok=True)
    df1.to_excel("tests/fixtures/test_out_e1.xlsx", index=False)
    df2.to_excel("tests/fixtures/test_out_e2.xlsx", index=False)
    wb.save("tests/fixtures/test_out_e3.xlsx")
    
    with open("tests/fixtures/test_out_e1.xlsx", "rb") as f:
        client.post(f"/api/sessions/{s_id}/files/excel1", files={"file": ("test_out_e1.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    with open("tests/fixtures/test_out_e2.xlsx", "rb") as f:
        client.post(f"/api/sessions/{s_id}/files/excel2", files={"file": ("test_out_e2.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    with open("tests/fixtures/test_out_e3.xlsx", "rb") as f:
        client.post(f"/api/sessions/{s_id}/files/excel3", files={"file": ("test_out_e3.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        
    mapping = {
        "excel1": {"sheet": "Sheet1", "columns": {"job_number": "Job No.", "balance_quantity": "Balance Quantity", "ocs_date": "OCS Date"}},
        "excel2": {"sheet": "Sheet1", "columns": {"job_number": "JOB NO.", "inspection_from": "FROM", "inspection_upto": "UPTO", "date_received": "DATE RECEIVED", "qap_appl": "QAP APPL.", "no_of_working_days": "NO. OF DAYS"}},
        "excel3": {"sheet": "Sheet1", "columns": {"job_number": "Job No.", "running_orders": "Running Orders", "orders_for_fd": "Orders for FD", "ocs_done": "OCS done", "expediting": "Exp.", "inspection": "Inspn", "others": "Others", "total": "Total"}}
    }
    client.post(f"/api/sessions/{s_id}/mapping", json=mapping)
    
    # 4. Generate Combined and approve all
    job_numbers = ["B269", "B224", "B285", "B302", "B378"]
    client.post(f"/api/sessions/{s_id}/calculations/combined", json={"job_numbers": job_numbers, "evaluation_month": "2026-08"})
    
    for job in job_numbers:
        client.post(f"/api/sessions/{s_id}/jobs/{job}/approve", json={"acknowledge_warnings": True})
        
    # 5. Generate Output
    r_out = client.post(f"/api/sessions/{s_id}/output/generate", json={"job_numbers": job_numbers, "evaluation_month": "2026-08"})
    assert r_out.status_code == 200
    
    out_path = r_out.json()["output_path"]
    
    # 6. Verify Output Excel file directly
    out_wb = openpyxl.load_workbook(out_path, data_only=True)
    out_ws = out_wb.active
    
    # Check regex replacement
    assert out_ws.cell(row=1, column=1).value == "Consolidated man hour requirement for Aug'26"
    assert out_ws.cell(row=2, column=5).value == "MH to be released for Aug'26"
    
    # Map column headers to index
    header_row = 3
    headers_map = {out_ws.cell(row=header_row, column=c).value: c for c in range(1, out_ws.max_column + 1) if out_ws.cell(row=header_row, column=c).value}
    
    job_col = headers_map["Job No."]
    insp_col = headers_map["Inspn"]
    others_col = headers_map["Others"]
    total_col = headers_map["Total"]
    
    results = {}
    for r in range(header_row + 1, out_ws.max_row + 1):
        j = out_ws.cell(row=r, column=job_col).value
        if j and j not in ["Leave", "Total"]:
            results[j] = {
                "inspn": out_ws.cell(row=r, column=insp_col).value,
                "others": out_ws.cell(row=r, column=others_col).value,
                "total": out_ws.cell(row=r, column=total_col).value
            }
            
    assert results["B224"]["others"] == 12.5
    assert results["B269"]["others"] == 0.5
    assert results["B285"]["others"] == 2
    assert results["B302"]["others"] == 3.5
    assert results["B378"]["others"] == 7.25

    # Test month 2026-09 logic
    r_out_sep = client.post(f"/api/sessions/{s_id}/output/generate", json={"job_numbers": job_numbers, "evaluation_month": "2026-09"})
    assert r_out_sep.status_code == 200
    sep_path = r_out_sep.json()["output_path"]
    sep_wb = openpyxl.load_workbook(sep_path, data_only=True)
    sep_ws = sep_wb.active
    assert sep_ws.cell(row=1, column=1).value == "Consolidated man hour requirement for Sep'26"
