import pytest
from app.services.output_engine import OutputEngine
from app.schemas.output import OutputGenerateRequest, CustomColumnData
from app.services.session_service import SessionService
from app.schemas.session import Session
from app.schemas.mapping import MappingConfiguration, Excel3Mapping, Excel3MappingFields, Excel2Mapping, Excel2MappingFields, Excel1Mapping, Excel1MappingFields
import os
import uuid
import openpyxl
from unittest.mock import patch
from app.schemas.review import JobReviewResult

def mock_get_reviews(s, j, e):
    return [
        JobReviewResult(
            job_number="B123",
            status="APPROVED",
            warnings=[],
            evidence=[],
            overrides={},
            calculated_total=10,
            fd=1, running_orders=2, ocs_done=3, expediting=4, inspection=5, others=6, meeting=7,
            native_expediting_used=False
        )
    ]

@patch('app.services.output_engine.ReviewService.get_reviews', side_effect=mock_get_reviews)
def test_repeated_generation(mock_reviews):
    session_id = str(uuid.uuid4())
    session = SessionService.create_session()
    session.session_id = session_id
    
    from app.config import UPLOAD_DIR
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    e3_id = "test_e3_id"
    e3_path = os.path.join(UPLOAD_DIR, f"{e3_id}.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Make a merged cell for Total
    ws.cell(row=3, column=1, value="Job No")
    ws.cell(row=3, column=2, value="Running Orders")
    ws.cell(row=3, column=3, value="Total")
    ws.merge_cells("C3:E3")  # Merged Total (spans C to E)
    ws.cell(row=4, column=1, value="B123")
    
    wb.save(e3_path)
    
    session.excel3_file_id = e3_id
    session.mapping = MappingConfiguration(
        excel1=Excel1Mapping(sheet="Sheet1", columns=Excel1MappingFields(job_number="Job No", balance_quantity="Balance", ocs_date="OCS")),
        excel2=Excel2Mapping(sheet="Sheet1", columns=Excel2MappingFields(job_number="Job No", inspection_from="From", inspection_upto="Upto")),
        excel3=Excel3Mapping(
            sheet="Sheet1",
            columns=Excel3MappingFields(
                job_number="Job No",
                running_orders="Running Orders",
                total="Total",
                expediting="Exp",
                inspection="Insp",
                others="Others"
            )
        )
    )
    import app.services.session_service
    app.services.session_service._sessions[session_id] = session
    
    request1 = OutputGenerateRequest(
        job_numbers=["B123"],
        evaluation_month="2026-06",
        custom_columns=[]
    )
    
    meta1 = OutputEngine.generate_output(session_id, request1)
    
    request2 = OutputGenerateRequest(
        job_numbers=["B123"],
        evaluation_month="2026-06",
        custom_columns=[]
    )
    
    meta2 = OutputEngine.generate_output(session_id, request2)
    
    assert meta1.original_sha256 == meta2.original_sha256
    
    wb1 = openpyxl.load_workbook(meta1.output_path)
    wb2 = openpyxl.load_workbook(meta2.output_path)
    
    ws1 = wb1.active
    ws2 = wb2.active
    
    assert ws1.max_column == ws2.max_column
    assert len(ws1.merged_cells.ranges) == len(ws2.merged_cells.ranges)
