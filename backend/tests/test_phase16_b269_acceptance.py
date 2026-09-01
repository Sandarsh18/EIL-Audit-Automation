import pytest
from fastapi.testclient import TestClient
import pandas as pd
import os
import openpyxl
from app.main import app
from app.schemas.mapping import MappingConfiguration, Excel1Mapping, Excel1MappingFields, Excel2Mapping, Excel2MappingFields, Excel3Mapping, Excel3MappingFields

client = TestClient(app)

E3_REAL_PATH = "/home/1RV24MC093_SANDARSH_J_N/projects/eil1/backend/tests/fixtures/CONSOLIDATED ManhourAp25.xlsx"
os.makedirs("tests/fixtures", exist_ok=True)
if not os.path.exists(E3_REAL_PATH):
    # Create a dummy Excel 3 template to avoid skipping the test if not present
    df = pd.DataFrame({"Job No": ["B269"], "No. of Running orders": [999], "Orders for FD f/up": [888], "OCS done": [777], "Exp.": [666], "Inspn": [555], "Others": [444], "Total": [333], "*Others/ Remarks": [None], "Insp/ Project Coordinators": [None], "MH available as on": [None], "MH to be released for Aug'26": [None], "Allotted": [None], "8518 reqd": [None]})
    df.to_excel(E3_REAL_PATH, sheet_name="ConsolidatedMHrequirementFeb26", index=False)

def test_phase16_b269_acceptance():
    resp = client.post("/api/sessions")
    s_id = resp.json()["session_id"]
    
    # 1. B269 dataset for Excel 1 (Testing FD, Running Orders, OCS Done)
    # Target: Running Orders=4, FD=1, OCS Done=9
    # Record 1: Running Order (Balance != 0)
    # Record 2: Running Order
    # Record 3: Running Order
    # Record 4: Running Order
    # Record 5: FD (Balance == 0, OCS Blank)
    # Record 6-14: OCS Done (Balance == 0, OCS populated)
    bqs = [10, 20, 30, 40, 0] + [0]*9
    ocs = [None]*5 + ["01/08/2026"]*9
    df1 = pd.DataFrame({
        "Job No.": ["B269"] * 14,
        "Balance Quantity": bqs,
        "OCS Date": ocs
    })
    
    # 2. Excel 2 Inspection Data with strict month bounds and cross-month records
    df2 = pd.DataFrame({
        "Job No.": ["B269"] * 9,
        "Inspection Attended (From)": [
            "2025-08-01", # Wrong Year
            "2026-07-28", # Wrong Month
            "2026-08-01", # VALID (1 day)
            "2026-08-15", # VALID (3 days)
            "2026-08-30", # VALID (1 day)
            "2026-09-01", # Wrong Month
            "2027-08-01", # Wrong Year
            "2026-07-31", # CROSS MONTH (Excluded)
            "2026-08-30"  # CROSS MONTH (Excluded)
        ],
        "Inspection Attended (Upto)": [
            "2025-08-02",
            "2026-07-30",
            "2026-08-02",
            "2026-08-18",
            "2026-08-31",
            "2026-09-02",
            "2027-08-02",
            "2026-08-01",
            "2026-09-02"
        ],
        "No. of Days": [1, 2, 1, 3, 1, 1, 1, 2, 3],
    })
    
    # Total valid = 1 + 3 + 1 = 5 days.
    
    df1.to_excel("tests/fixtures/real_e1.xlsx", index=False)
    df2.to_excel("tests/fixtures/real_e2.xlsx", index=False)
    
    with open("tests/fixtures/real_e1.xlsx", "rb") as f:
        client.post(f"/api/sessions/{s_id}/files/excel1", files={"file": ("real_e1.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    with open("tests/fixtures/real_e2.xlsx", "rb") as f:
        client.post(f"/api/sessions/{s_id}/files/excel2", files={"file": ("real_e2.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    with open(E3_REAL_PATH, "rb") as f:
        client.post(f"/api/sessions/{s_id}/files/excel3", files={"file": ("CONSOLIDATED ManhourAp25.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        
    mapping = MappingConfiguration(
        excel1=Excel1Mapping(sheet="Sheet1", columns=Excel1MappingFields(job_number="Job No.", balance_quantity="Balance Quantity", ocs_date="OCS Date")),
        excel2=Excel2Mapping(sheet="Sheet1", columns=Excel2MappingFields(job_number="Job No.", inspection_from="Inspection Attended (From)", inspection_upto="Inspection Attended (Upto)")),
        excel3=Excel3Mapping(sheet="ConsolidatedMHrequirementFeb26", columns=Excel3MappingFields(
            job_number="Job No", running_orders="No. of Running orders", orders_for_fd="Orders for FD f/up", ocs_done="OCS done",
            expediting="Exp.", inspection="Inspn", others="Others", total="Total"
        ))
    )
    from app.services.session_service import SessionService
    SessionService.update_session_mapping(s_id, mapping)
    client.post(f"/api/sessions/{s_id}/evaluation-month", json={"evaluation_month": "2026-08"})
    
    # 3. Calculate Combined
    resp_calc = client.post(f"/api/sessions/{s_id}/calculations/combined", json={"job_numbers": ["B269"], "evaluation_month": "2026-08"})
    results = resp_calc.json()
    
    assert len(results) == 1
    b269 = results[0]
    
    assert b269["fd"] == 1
    assert b269["running_orders"] == 4
    assert b269["ocs_done"] == 9
    assert b269["expediting"] == 26 # (4 + 9) * 2
    assert b269["inspection"] == 24.0 # 3 valid records * 8
    assert b269["others"] == 0
    assert b269["calculated_total"] == 50.0 # 26 + 24 + 0
    
    # 4. Generate Output (Approve and write)
    client.post(f"/api/sessions/{s_id}/jobs/B269/approve", json={"acknowledge_warnings": True})
    
    # Generate Output Plan
    req_gen = {
        "job_numbers": ["B269"],
        "resolved_destinations": {},
        "approved_formula_overwrites": [],
    }
    meta = client.post(f"/api/sessions/{s_id}/output/generate", json=req_gen).json()
    
    wb_out = openpyxl.load_workbook(meta["output_path"])
    sheet = wb_out.active
    
    # Verify the EXACT 14 columns
    EXPECTED_HEADERS = [
        "Job No", "No. of Running orders", "Orders for FD f/up", "OCS done",
        "Exp.", "Inspn", "Others", "Meeting", "Total", "*Others/ Remarks",
        "Insp/ Project Coordinators", "MH available as on", 
        "MH to be released for Aug'26", "Allotted", "8518 reqd"
    ]
    
    header_vals = []
    for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
        if any(row):
            header_vals = [str(v).strip() for v in row if v is not None]
            break
            
    assert header_vals == EXPECTED_HEADERS
    
    # Check if contaminated Excel 3 defaults were ignored
    # We injected 999, 888, 777 etc. in the dummy Excel 3 template.
    # The output should NOT have those values. It should have the calculated ones.
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip().lower() == "b269":
            assert row[1] == 4 # Running Orders
            assert row[2] == 1 # FD (Orders for FD)
            assert row[3] == 9 # OCS
            assert row[4] == 26 # Exp
            assert row[5] == 24 # Insp (3 records * 8)
            assert row[6] == 0 # Others
            assert row[7] == 0 # Meeting (added in previous commit)
            assert row[8] == 50 # Total (26 + 24 + 0)
            break
            
    # Also verify inspection evidence payload directly to ensure we have exactly 3 valid and 6 invalid/excluded records
    resp_insp = client.post(f"/api/sessions/{s_id}/calculations/inspection", json={"job_numbers": ["B269"], "evaluation_month": "2026-08"})
    insp_data = resp_insp.json()[0]
    
    assert insp_data["records_analyzed"] == 9
    assert insp_data["valid_records"] == 3
    assert insp_data["invalid_records"] == 0 # None were structurally invalid, but 6 were excluded
    assert insp_data["evaluation_month_str"] == "2026-08"
    
    # Verify the 6 excluded records were logged in evidence with "EXCLUDED" status
    excluded = [e for e in insp_data["evidence"] if e["status"] == "EXCLUDED"]
    assert len(excluded) == 6
