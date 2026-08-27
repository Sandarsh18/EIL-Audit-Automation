import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

def create_synthetic():
    wb = openpyxl.Workbook()
    
    # Sheet 1: Unrelated Sheet
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "This is a summary sheet"
    ws1["A1"].font = Font(bold=True)
    
    # Sheet 2: Manhour (The target sheet)
    ws2 = wb.create_sheet("Manhour")
    
    headers = ["Job No.", "Supplier", "Description", "FD", "Running Orders", "OCS Done", "Expediting", "Inspection", "Others", "Total"]
    for col_idx, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col_idx, value=header)
        
    # Row 2: B269 (Unique)
    ws2.append(["B269", "ABC", "Pump", 0, 2, 2, 8, 16, 5, 29])
    
    # Row 3: B378 (Unique)
    ws2.append(["B378", "XYZ", "Valve", 1, 1, 3, 8, 8, 0, 16])
    
    # Row 4: B440 (Ambiguous - match 1)
    ws2.append(["B440", "ABC", "Compressor", 1, 5, 2, 14, 24, 0, 38])
    
    # Row 5: B440 (Ambiguous - match 2)
    ws2.append(["B440", "XYZ", "Motor", 0, 3, 1, 8, 16, 0, 24])
    
    # Row 6: B550 (Formula Target)
    ws2.append(["B550", "QWE", "Pipe", 1, 2, 2, 8, 16, 0, 0])
    ws2.cell(row=6, column=10).value = "=SUM(G6:I6)" # Total column formula
    
    # Formatting
    ws2["A2"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Merged cells (unrelated to target data, e.g. a title)
    ws2.merge_cells("A15:C15")
    ws2["A15"] = "Merged Area"
    
    # Sheet 3: Hidden Sheet
    ws3 = wb.create_sheet("HiddenSheet")
    ws3.sheet_state = 'hidden'
    ws3["A1"] = "Hidden data"
    
    os.makedirs("backend/tests/fixtures", exist_ok=True)
    wb.save("backend/tests/fixtures/synthetic_excel3.xlsx")
    
if __name__ == "__main__":
    create_synthetic()
