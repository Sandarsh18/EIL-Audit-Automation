import openpyxl

wb = openpyxl.load_workbook("local_test_data/CONSOLIDATED ManhourAp25.xlsx", data_only=False)
sheet = wb["ConsolidatedMHrequirementApr25"]

print("==== WORKBOOK STRUCTURE ====")
print(f"Sheet: ConsolidatedMHrequirementApr25")

# Find headers
h_row = None
job_col = None
exp_col = None
tot_col = None

for r in range(1, 10):
    for c in range(1, 20):
        val = str(sheet.cell(row=r, column=c).value).strip() if sheet.cell(row=r, column=c).value else ""
        if "Job No" in val or val == "Job":
            h_row = r
            job_col = c
        if val == "Exp.":
            exp_col = c
        if val == "Total":
            tot_col = c
            
print(f"Header row: {h_row}")
from openpyxl.utils import get_column_letter
print(f"Job Column: {get_column_letter(job_col)} (Logical: job_number)")
print(f"Exp. Column: {get_column_letter(exp_col)} (Logical: expediting)")
print(f"Total Column: {get_column_letter(tot_col)} (Logical: total)")

print("\n==== EXPEDITING FORMULA INVENTORY ====")
formulas = 0
hardcoded = 0
blanks = 0

# Find max row
max_row = sheet.max_row
for row in range(h_row + 1, max_row + 1):
    job = sheet.cell(row=row, column=job_col).value
    if not job or str(job).strip() == "":
        continue
        
    val = sheet.cell(row=row, column=exp_col).value
    
    if val is None:
        blanks += 1
        # print(f"Row {row}: Job {job} -> BLANK")
    elif isinstance(val, str) and val.startswith("="):
        formulas += 1
        print(f"Row {row}: Job {job} -> FORMULA: {val}")
    else:
        hardcoded += 1
        # print(f"Row {row}: Job {job} -> HARDCODED: {val}")
        
print(f"\nExpediting Summary: Formulas={formulas}, Hardcoded={hardcoded}, Blanks={blanks}")

print("\n==== TOTAL FORMULA INVENTORY ====")
t_formulas = 0
t_hardcoded = 0
t_blanks = 0

for row in range(h_row + 1, max_row + 1):
    job = sheet.cell(row=row, column=job_col).value
    if not job or str(job).strip() == "":
        continue
        
    val = sheet.cell(row=row, column=tot_col).value
    if val is None:
        t_blanks += 1
    elif isinstance(val, str) and val.startswith("="):
        t_formulas += 1
        print(f"Row {row}: Job {job} -> FORMULA: {val}")
    else:
        t_hardcoded += 1

print(f"\nTotal Summary: Formulas={t_formulas}, Hardcoded={t_hardcoded}, Blanks={t_blanks}")
